from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from bson import ObjectId
from io import BytesIO
from database import db
from helpers import serialize_doc, ifilter, hash_password, notify_parent, FRONTEND_URL
from dependencies import require_admin
from models import ParentInvite
import uuid
import openpyxl

admin_router = APIRouter(prefix="/admin", tags=["admin"])


@admin_router.get("/fee-queries")
async def get_all_fee_queries(
    user: dict = Depends(require_admin),
    page: int = 1, limit: int = 20,
    status: str = None,
):
    import math
    q = {}
    if status and status != "all":
        q["status"] = status
    total = await db.fee_queries.count_documents(q)
    skip = (max(1, page) - 1) * limit
    queries = await db.fee_queries.find(q).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {
        "items": [serialize_doc(q) for q in queries],
        "total": total, "page": page,
        "pages": max(1, math.ceil(total / limit))
    }


@admin_router.patch("/fee-queries/{query_id}/comment")
async def add_fee_query_comment(query_id: str, data: dict, user: dict = Depends(require_admin)):
    comment = (data.get("admin_comment") or "").strip()
    if not comment:
        raise HTTPException(status_code=400, detail="Comment cannot be empty")
    await db.fee_queries.update_one(
        {"_id": ObjectId(query_id)},
        {"$set": {"admin_comment": comment, "commented_at": datetime.now(timezone.utc),
                  "commented_by": user.get("name", "Admin")}}
    )
    return serialize_doc(await db.fee_queries.find_one({"_id": ObjectId(query_id)}))


@admin_router.patch("/fee-queries/{query_id}/resolve")
async def resolve_fee_query(query_id: str, user: dict = Depends(require_admin)):
    await db.fee_queries.update_one(
        {"_id": ObjectId(query_id)},
        {"$set": {"status": "resolved", "resolved_at": datetime.now(timezone.utc),
                  "resolved_by": user.get("id") or str(user.get("_id", ""))}}
    )
    query = await db.fee_queries.find_one({"_id": ObjectId(query_id)})
    if not query:
        raise HTTPException(status_code=404, detail="Query not found")
    return serialize_doc(query)


@admin_router.get("/backup")
async def download_backup(user: dict = Depends(require_admin)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enquiries"
    ws.append(["ID", "Name", "Email", "Phone", "City", "Source", "Stage", "Notes", "Created At"])
    for e in await db.enquiries.find().to_list(10000):
        ws.append([str(e.get("_id")), e.get("student_name",""), e.get("email",""), e.get("phone",""),
                   e.get("city",""), e.get("source",""), e.get("stage",""), e.get("notes",""), str(e.get("created_at",""))])
    ws2 = wb.create_sheet("Students")
    ws2.append(["ID", "Name", "Email", "Phone", "Branch", "Status", "DOB", "Guardian Name", "Guardian Phone", "ID Proof", "Institute", "Enrollment Date"])
    for s in await db.students.find().to_list(10000):
        ws2.append([str(s.get("_id")), s.get("name",""), s.get("email",""), s.get("phone",""),
                    s.get("branch_id",""), s.get("status",""), s.get("dob",""), s.get("guardian_name",""),
                    s.get("guardian_phone",""), s.get("id_proof",""), s.get("institute_name",""), str(s.get("enrollment_date",""))])
    ws3 = wb.create_sheet("Invoices")
    ws3.append(["ID", "Student Name", "Course Name", "Base Fee", "GST", "Discount", "Total", "Paid", "Balance", "Status"])
    for inv in await db.invoices.find().to_list(10000):
        ws3.append([str(inv.get("_id")), inv.get("student_name",""), inv.get("course_name",""),
                    inv.get("base_fee",0), inv.get("gst_amount",0), inv.get("discount",0),
                    inv.get("total",0), inv.get("paid_amount",0), inv.get("balance",0), inv.get("status","")])
    ws4 = wb.create_sheet("Attendance")
    ws4.append(["Session ID", "Student ID", "Student Name", "Status", "Method", "Date"])
    for a in await db.attendance.find().to_list(50000):
        ws4.append([a.get("session_id",""), a.get("student_id",""), a.get("student_name",""),
                    a.get("status",""), a.get("method","manual"), str(a.get("created_at",""))])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=edutech_backup_{ts}.xlsx"})


@admin_router.delete("/data")
async def delete_all_data(data: dict, user: dict = Depends(require_admin)):
    if data.get("confirm") != "DELETE ALL":
        raise HTTPException(status_code=400, detail='Confirmation required: send {"confirm": "DELETE ALL"}')
    counts = {
        "enquiries":  (await db.enquiries.delete_many({})).deleted_count,
        "students":   (await db.students.delete_many({})).deleted_count,
        "invoices":   (await db.invoices.delete_many({})).deleted_count,
        "payments":   (await db.payments.delete_many({})).deleted_count,
        "fee_queries":(await db.fee_queries.delete_many({})).deleted_count,
        "attendance": (await db.attendance.delete_many({})).deleted_count,
    }
    return {"deleted": counts, "message": "All CRM and student data has been deleted"}


@admin_router.post("/restore")
async def restore_from_backup(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid .xlsx file")
    results = {}
    if "Enquiries" in wb.sheetnames:
        ws = wb["Enquiries"]
        hdrs = [c.value for c in ws[1]]
        inserted = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(hdrs, row))
            if not d.get("Email"):
                continue
            if not await db.enquiries.find_one({"email": str(d.get("Email","")).lower()}):
                await db.enquiries.insert_one({
                    "student_name": d.get("Name",""), "email": str(d.get("Email","")).lower(),
                    "phone": str(d.get("Phone","")), "city": d.get("City",""),
                    "source": d.get("Source","manual"), "stage": d.get("Stage","new"),
                    "notes": d.get("Notes",""), "created_at": datetime.now(timezone.utc),
                })
                inserted += 1
        results["enquiries"] = inserted
    if "Students" in wb.sheetnames:
        ws = wb["Students"]
        hdrs = [c.value for c in ws[1]]
        inserted = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(hdrs, row))
            if not d.get("Email"):
                continue
            if not await db.students.find_one({"email": str(d.get("Email","")).lower()}):
                await db.students.insert_one({
                    "name": d.get("Name",""), "email": str(d.get("Email","")).lower(),
                    "phone": str(d.get("Phone","")), "branch_id": d.get("Branch",""),
                    "status": d.get("Status","onboarding"), "dob": d.get("DOB",""),
                    "guardian_name": d.get("Guardian Name",""), "guardian_phone": str(d.get("Guardian Phone","")),
                    "id_proof": d.get("ID Proof",""), "institute_name": d.get("Institute",""),
                    "created_at": datetime.now(timezone.utc),
                })
                inserted += 1
        results["students"] = inserted
    return {"restored": results, "message": f"Restore complete: {results}"}


# Parent management (admin)
@admin_router.get("/parents")
async def list_parents(user: dict = Depends(require_admin)):
    parents = await db.users.find(ifilter(user, {"role": "parent"}), {"password_hash": 0}).to_list(1000)
    result = []
    for p in parents:
        pd = serialize_doc(p)
        if p.get("student_id"):
            try:
                student = await db.students.find_one({"_id": ObjectId(p["student_id"])})
                pd["student_name"] = (student or {}).get("name", "Unknown")
                pd["student_email"] = (student or {}).get("email", "")
            except Exception:
                pd["student_name"] = "Unknown"
        result.append(pd)
    return result


@admin_router.post("/parents")
async def create_parent_account(data: ParentInvite, user: dict = Depends(require_admin)):
    email = data.parent_email.lower().strip()
    iid = user.get("institute_id")
    if await db.users.find_one({"email": email, "institute_id": iid} if iid else {"email": email}):
        raise HTTPException(status_code=400, detail="Email already registered")
    student = None
    try:
        student = await db.students.find_one({"_id": ObjectId(data.student_id)})
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid student ID")
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    tmp_password = f"Parent@{uuid.uuid4().hex[:6]}"
    parent_doc = {
        "name": data.parent_name.strip(),
        "email": email,
        "password_hash": hash_password(tmp_password),
        "role": "parent",
        "student_id": data.student_id,
        "phone": data.parent_phone or "",
        "institute_id": iid,
        "created_at": datetime.now(timezone.utc),
    }
    result = await db.users.insert_one(parent_doc)
    await db.students.update_one(
        {"_id": ObjectId(data.student_id)},
        {"$addToSet": {"parent_ids": str(result.inserted_id)}}
    )
    html = f"""<html><body style="font-family:sans-serif;max-width:600px;margin:0 auto">
    <div style="background:#002EB8;padding:24px 28px;border-radius:8px 8px 0 0">
      <h1 style="color:white;margin:0;font-size:20px">EduTech LMS — Parent Portal Access</h1>
    </div>
    <div style="background:#F8F9FA;border:1px solid #E5E7EB;border-radius:0 0 8px 8px;padding:28px">
      <p>Dear <strong>{data.parent_name}</strong>,</p>
      <p>You have been granted access to the parent portal for <strong>{student.get('name')}</strong>.</p>
      <p>Login at: <a href="{FRONTEND_URL}/parent-portal">{FRONTEND_URL}/parent-portal</a></p>
      <p><strong>Email:</strong> {email}<br><strong>Temporary Password:</strong> {tmp_password}</p>
      <p style="color:#FF2B2B;font-size:12px">Please change your password after first login.</p>
    </div></body></html>"""
    await notify_parent(email, data.parent_name, student.get("name", ""), "EduTech LMS — Parent Portal Access", html)
    return {"id": str(result.inserted_id), "email": email, "temp_password": tmp_password, "student_name": student.get("name")}


@admin_router.delete("/parents/{parent_id}")
async def delete_parent(parent_id: str, user: dict = Depends(require_admin)):
    parent = await db.users.find_one({"_id": ObjectId(parent_id)})
    if parent and parent.get("student_id"):
        await db.students.update_one(
            {"_id": ObjectId(parent["student_id"])},
            {"$pull": {"parent_ids": parent_id}}
        )
    await db.users.delete_one({"_id": ObjectId(parent_id)})
    return {"message": "Parent account deleted"}
